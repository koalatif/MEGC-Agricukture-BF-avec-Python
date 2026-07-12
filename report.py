"""Reporting du MEGC BF: agrégats macro, bien-être (variation équivalente), revenus,
tableaux sectoriels, export Excel et graphiques du sentier dynamique."""
import numpy as np, os
np.seterr(all='ignore')

def _labels(m):
    """Libellés de branches : cherche une MCS xlsx à côté du modèle (chemin portable)."""
    import glob
    here=os.path.dirname(os.path.abspath(__file__))
    for p in glob.glob(os.path.join(here,"*.xlsx"))+glob.glob(os.path.join(here,"..","*.xlsx")):
        if "MCS" in os.path.basename(p):
            try:
                import openpyxl
                ws=openpyxl.load_workbook(p,data_only=True).active
                return {str(ws.cell(r,2).value).strip():str(ws.cell(r,1).value).strip()
                        for r in range(4,ws.max_row+1) if ws.cell(r,2).value}
            except Exception: pass
    return {}

def equivalent_variation(m, r0, r1):
    """Variation équivalente par ménage (LES/Stone-Geary), en milliards FCFA."""
    PC0=r0['PC']; gam=m.gam; Cmin=m.Cmin
    ev={}
    for h,name in enumerate(m.d.H):
        u1=r1['util'][h]
        # e(p0,u1) = sum p0 Cmin + exp(u1) * prod (p0/gam)^gam
        term=np.exp(u1)*np.prod(np.where(gam[:,h]>0,(PC0/np.where(gam[:,h]>0,gam[:,h],1))**gam[:,h],1.0))
        e_p0_u1=(PC0*Cmin[:,h]).sum()+term
        Y0=r0['CTH'][h] if 'CTH' in r0 else (PC0*r0['C'][:,h]).sum()
        ev[name]=(e_p0_u1-Y0)/1e3
    return ev

def macro_table(m, r0, r1=None):
    def g(r): return dict(PIB_VA=r['GDP_VA']/1e3, Conso=r['Conso']/1e3, Invest=r['Invest']/1e3,
        Gov=r['Gov']/1e3, Export=r['Export']/1e3, Import=r['Import']/1e3,
        Recettes_pub=r['GVTrev']/1e3, Epargne_pub=r.get('SG',0.0)/1e3,
        Salaire_moy=float(np.mean(r['W'])), IPC=float(r.get('cpi',1.0)))
    base=g(r0)
    if r1 is None: return base
    sim=g(r1); return {k:(base[k],sim[k],100*(sim[k]-base[k])/base[k] if base[k] else 0) for k in base}

def sectoral_table(m, r0, r1=None, top=15):
    lab=_labels(m); J=m.d.J
    va0=r0['VA']; xst0=r0['XST']; ld0=r0['LD'].sum(0)
    rows=[]
    for j in range(len(J)):
        row=[J[j], lab.get(J[j],J[j])[:34], va0[j]/1e3]
        if r1 is not None:
            row+= [100*(r1['XST'][j]-xst0[j])/xst0[j] if xst0[j] else 0,
                   100*(r1['VA'][j]-va0[j])/va0[j] if va0[j] else 0,
                   100*(r1['PVA'][j]-r0['PVA'][j])/r0['PVA'][j] if r0['PVA'][j] else 0]
        rows.append(row)
    rows.sort(key=lambda x:-x[2])
    return rows[:top]

def income_table(m, r0, r1=None):
    out={}
    for h,name in enumerate(m.d.H):
        rec=dict(revenu=r0['YH'][h]/1e3, disponible=r0['YDH'][h]/1e3, epargne=r0['SH'][h]/1e3)
        if r1 is not None:
            rec['var_revenu_%']=100*(r1['YH'][h]-r0['YH'][h])/r0['YH'][h] if r0['YH'][h] else 0
        out[name]=rec
    return out

def print_summary(m, r0, r1=None, title="Résultats"):
    print(f"\n===== {title} =====")
    mt=macro_table(m,r0,r1)
    if r1 is None:
        for k,v in mt.items(): print(f"  {k:14s}: {v:,.0f}")
    else:
        print(f"  {'Agrégat':16s}{'BAU':>12s}{'Choc':>12s}{'Var %':>9s}")
        for k,(b,s,d) in mt.items(): print(f"  {k:16s}{b:12,.0f}{s:12,.0f}{d:8.2f}%")
        ev=equivalent_variation(m,r0,r1)
        print("  Bien-être (variation équivalente, Mds FCFA):")
        for h,v in ev.items(): print(f"     {h:6s}: {v:+.1f}")

def export_excel(path, m, r0, r1=None, dyn_path=None, dyn_shock=None):
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb=openpyxl.Workbook(); HF=Font(bold=True,color='FFFFFF'); FILL=PatternFill('solid',fgColor='2F5496')
    def head(ws,cols,row=1):
        for j,t in enumerate(cols):
            c=ws.cell(row,j+1,t); c.font=HF; c.fill=FILL
    ws=wb.active; ws.title="Macro"; head(ws,["Agrégat (Mds FCFA)","BAU","Choc","Var %"])
    mt=macro_table(m,r0,r1)
    if r1 is None:
        for i,(k,v) in enumerate(mt.items()): ws.cell(i+2,1,k); ws.cell(i+2,2,round(v,1))
    else:
        for i,(k,(b,s,d)) in enumerate(mt.items()):
            ws.cell(i+2,1,k); ws.cell(i+2,2,round(b,1)); ws.cell(i+2,3,round(s,1)); ws.cell(i+2,4,round(d,2))
        ws2=wb.create_sheet("Bien-être"); head(ws2,["Ménage","Var. équivalente (Mds)"])
        for i,(h,v) in enumerate(equivalent_variation(m,r0,r1).items()):
            ws2.cell(i+2,1,h); ws2.cell(i+2,2,round(v,2))
    wsi=wb.create_sheet("Revenus"); head(wsi,["Ménage","Revenu","Disponible","Épargne","Var revenu %"])
    for i,(h,rec) in enumerate(income_table(m,r0,r1).items()):
        wsi.cell(i+2,1,h); wsi.cell(i+2,2,round(rec['revenu'],1)); wsi.cell(i+2,3,round(rec['disponible'],1))
        wsi.cell(i+2,4,round(rec['epargne'],1))
        if 'var_revenu_%' in rec: wsi.cell(i+2,5,round(rec['var_revenu_%'],2))
    wss=wb.create_sheet("Secteurs"); 
    cols=["Code","Branche","VA (Mds)"]+(["ΔProd %","ΔVA %","ΔPrix VA %"] if r1 is not None else [])
    head(wss,cols)
    for i,row in enumerate(sectoral_table(m,r0,r1,top=88)):
        for j,v in enumerate(row): wss.cell(i+2,j+1, round(v,2) if isinstance(v,float) else v)
    if dyn_path is not None:
        wsd=wb.create_sheet("Dynamique"); head(wsd,["Période","PIB(VA) BAU"]+(["PIB(VA) Choc"] if dyn_shock else []))
        for t,p in enumerate(dyn_path):
            wsd.cell(t+2,1,t); wsd.cell(t+2,2,round(p['GDP_VA']/1e3,1))
            if dyn_shock: wsd.cell(t+2,3,round(dyn_shock[t]['GDP_VA']/1e3,1))
    wb.save(path); return path

def plot_dynamic(path_bau, path_shock=None, out="sentier_dynamique.png", labels=("BAU","Choc")):
    try:
        import matplotlib; matplotlib.use('Agg'); import matplotlib.pyplot as plt
    except Exception: return None
    t=[p['t'] for p in path_bau]; g=[p['GDP_VA']/1e3 for p in path_bau]
    fig,ax=plt.subplots(figsize=(7,4))
    ax.plot(t,g,'o-',label=labels[0])
    if path_shock is not None: ax.plot([p['t'] for p in path_shock],[p['GDP_VA']/1e3 for p in path_shock],'s--',label=labels[1])
    ax.set_xlabel("Période"); ax.set_ylabel("PIB (VA), Mds FCFA"); ax.set_title("Sentier dynamique du PIB")
    ax.legend(); ax.grid(alpha=0.3); fig.tight_layout(); fig.savefig(out,dpi=110); plt.close(fig)
    return out
